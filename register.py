from atexit import register
from cgitb import text
from tkinter import font
import pandas as pd
import openpyxl
import tkinter 
from tkinter import messagebox
from tkinter import simpledialog
from tkinter import*
from datetime import datetime
from tkinter import ttk



wb = openpyxl.load_workbook('MANTRA OEE ALMA ( AUTO)copy.xlsx', data_only=True)
ws = wb['Hourly Prod data ALMA Detail']



window=Tk()
window.title('testing')
window.geometry("900x600")



menubar = Menu(window)


filemenu = Menu(menubar, tearoff=0)
filemenu.add_command(label="New")
filemenu.add_separator()

filemenu.add_command(label="Exit", command=window.quit)
menubar.add_cascade(label="File", menu=filemenu)

editmenu = Menu(menubar, tearoff=0)
editmenu.add_command(label="Undo")

editmenu.add_separator()

menubar.add_cascade(label="GRNR")

menubar.add_cascade(label="CALIBRATION", menu=editmenu)
helpmenu = Menu(menubar, tearoff=0)
helpmenu.add_command(label="Help Index")
helpmenu.add_command(label="About...")
menubar.add_cascade(label="Help", menu=helpmenu)





label = tkinter.Label(window, text="CALIBRATION AND GRNR DIRECTORY", font=("bold", 20))
label.place(x=200)

labelver = tkinter.Label(window, text="Version 1.0.0", font=(10))
labelver.place(x=380,y=40)

labelauto = tkinter.Label(window, text="Auto Engineering Control", font=(6))
labelauto.place(x=300, y=70)



label_codeno=tkinter.Label(window,text="Code No : ",font=("bold",15))
label_codeno.place(x=200, y=100)

codeno = tkinter.Entry(window, width=50)
codeno.place(x=290, y=105)

label_warn = tkinter.Label(window, text="*SILA ISI MENGGUNAKAN HURUF BESAR", font=("bold", 12), fg="#ff0000")
label_warn.place(x=295, y=125)

top2 = Toplevel()
top2.geometry("1500x500")
top2.title('OVER DUE ITEM !!!')

result_codeno2 = tkinter.Label(top2, font=("bold", 18))
result_codeno2.place(x=10, y=100)

result_duedatein = tkinter.Label(top2, font=("bold", 18))
result_duedatein.place(x=230, y=100)

result_item = tkinter.Label(top2,font=("bold",18))
result_item.place(x=600, y=100)

result_location = tkinter.Label(top2,font=("bold",18))
result_location.place(x=1100, y=100)

result_adddue = tkinter.Label(top2, font=("bold", 18))
result_adddue.place(x=100, y=100)


def date():
    result_duedatein.config(text="")
    result_codeno2.config(text="")
    result_adddue.config(text="")
    result_item.config(text="")
    result_location.config(text="")

    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active

    due_date_str = ""
    code_no_str = ""
    item_str =""
    location_str=""
    Found = False
    for i in range(3, (ws.max_row)+1):
        if(isinstance(ws['I'+str(i)].value, datetime) and (ws['I'+str(i)].value-datetime.now()).days<0) :
            Found = True
            due_date_str += str(ws['I'+str(i)].value)+"\n"
            print(str(ws['F'+str(i)].value))
            c1 = my_sheet.cell(row=1, column=1)
            c1.value =(str(ws['F'+str(i)].value))
            
            code_no_str += str(ws['F'+str(i)].value)+"\n"
            print("okey")
            item_str += str(ws['B'+str(i)].value)+"\n"
            location_str += str(ws['J'+str(i)].value)+"\n"
            #break
        # if(ws['K'+str(i)].value)=='OVERDUE':
        #     Found = True
        #     print("okey")
        #     break
        
        # else:
        #     Found = False

    if (Found == True):
        result_duedatein.config(text="DUE DATE: \n"+due_date_str)
        result_codeno2.config(text="Code No:\n "+code_no_str)
        result_item.config(text= "Instrument : \n " +item_str)
        result_location.config(text="Location : \n " + location_str)
        # result_duedatein.config(text="DUE DATE: "+str(ws['I'+str(i)].value))
        # result_codeno2.config(text="Code No: "+str(ws['F'+str(i)].value))
        
    else:
        result_adddue.config(text="No record found.")


date()

def clear_text():
    codeno.delete(0, END)

def search():
    result_codeno.config(text="")
    result_inst.config(text="")
    result_model.config(text="")
    result_duedate.config(text="")
    result_location.config(text="")
    result_add.config(text="")
    result_vendor.config(text="")

    for i in range(2,(ws.max_row)+1):
        if((codeno.get()==ws['F'+str(i)].value)):
            Found=True
            break

        else:
            Found = False
            
            
    if (Found==True):
        result_codeno.config(text="Code No: "+str(ws['F'+str(i)].value))
        result_inst.config(text="Instrument: "+str(ws['B'+str(i)].value))
        result_model.config(text="Model No: "+str(ws['D'+str(i)].value))
        result_duedate.config(text="DUE DATE: "+str(ws['I'+str(i)].value))
        result_location.config(text="Location: "+str(ws['J'+str(i)].value))
        result_vendor.config(text="Vendor: "+str(ws['L'+str(i)].value))
    else:
        result_add.config(text="No record found.")       
    

def add():
    for i in range(2,(ws.max_row)+1):
        if((codeno.get() == ws['F'+str(i)].value)):
            Found=True
            break

        else:
            Found = False
              
    if(Found==True):
         messagebox.showinfo("Error", "Model Already Exists!")
                
    else:
        top1 = Toplevel()
        top1.geometry("500x800")
        top1.title('Calibration and GRNR Directory : Add Item')
        
        label_addtitle = tkinter.Label(top1, text="ADD ITEM", font=("bold",18))
        label_addtitle.place(x=190,y=10)

        label_arahan = tkinter.Label(top1, text="Sila Isikan Ruangan Di Bawah Dengan Lengkap", font=(14))
        label_arahan.place(x=3,y=50)
        
        label_instadd = tkinter.Label(top1, text="Instrument     :", font=("bold", 14))
        label_instadd.place(x=60,y= 100)
        instadd = tkinter.Entry(top1, width=40)
        instadd.place(x=200, y=100)
        
        label_manuadd = tkinter.Label(top1, text="Manufacturer  :", font=("bold", 14))
        label_manuadd.place(x=60, y=140)
        manuadd = tkinter.Entry(top1, width=40)
        manuadd.place(x=200, y=140)


        label_modeladd = tkinter.Label(top1, text="Model Number :",font=("bold", 14))
        label_modeladd.place(x=60,y=180)
        modeladd = tkinter.Entry(top1,width=40)
        modeladd.place(x=200,y=180)

        label_serialadd = tkinter.Label(top1, text="Serial Number :", font=("bold", 14))
        label_serialadd.place(x=60, y=220)
        serialadd = tkinter.Entry(top1,width=40)
        serialadd.place(x=200, y=220)

        label_codenoadd = tkinter.Label(top1, text="Code No        :",font=("bold", 14))
        label_codenoadd.place(x=60, y=260)
        codenoadd = tkinter.Entry(top1,width=40)
        codenoadd.place(x=200,y=260)
        label_warn1 = tkinter.Label(top1, text="*SILA ISI MENGGUNAKAN HURUF BESAR", font=("bold",9), fg="#ff0000")
        label_warn1.place(x=200, y=280)

        label_intervaladd = tkinter.Label(top1, text="Interval          :",font=("bold", 14))
        label_intervaladd.place(x=60,y=300) 
        intervaladd = tkinter.Entry(top1, width=40)
        intervaladd.place(x=200,y=300)


        label_caldateadd = tkinter.Label(top1, text="Cal Date        :", font=("bold", 14))
        label_caldateadd.place(x=60,y=340)
        caldateadd = tkinter.Entry(top1,width=40)
        caldateadd.place(x=200,y=340)

        label_dueadd = tkinter.Label(top1, text="Due Date       :",font=("bold", 14))
        label_dueadd.place(x=60,y=380)
        dueadd = tkinter.Entry(top1, width=40)
        dueadd.place(x=200,y=380)

        label_locationadd = tkinter.Label(top1, text="Location        :",font=("bold", 14))
        label_locationadd.place(x=60, y= 420)
        locationadd = tkinter.Entry(top1,width=40)
        locationadd.place(x=200, y=420)
        
        label_vendoradd = tkinter.Label(top1, text="Vendor          :",font=("bold", 14))
        label_vendoradd.place(x=60,y=500)
        vendoradd = tkinter.Entry(top1,width=40)
        vendoradd.place(x=200,y=500)
        
        label_treatmentadd = tkinter.Label(top1, text="Treatment          :",font=("bold", 14))
        label_treatmentadd.place(x=60,y=540)
        treatmentadd = tkinter.Entry(top1,width=40)
        treatmentadd.place(x=200,y=540)

        label_needadd = tkinter.Label(top1, text="Need Calibrate         :", font=("bold", 14))
        label_needadd.place(x=60, y=580)
        needadd = tkinter.Entry(top1, width=40)
        needadd.place(x=200, y=580)
        
        def clear_all():
            instadd.delete(0, END)
            manuadd.delete(0, END)
            modeladd.delete(0, END)
            serialadd.delete(0, END)
            codenoadd.delete(0, END)
            intervaladd.delete(0, END)
            caldateadd.delete(0, END)
            dueadd.delete(0, END)
            locationadd.delete(0, END)
            vendoradd.delete(0, END)
            treatmentadd.delete(0, END)
            needadd.delete(0, END)

        Button(top1, text="CLEAR", command=clear_all,
               font=("Helvetica", 11, "bold")).place(x=300, y=630)

        
        
        def added():
        
            res = messagebox.askyesno('Add', 'Record Succesfully Added! Do you want to add more?')
            if res == True:
                lastx = str((ws.max_row)+1)
                ws['B'+lastx] = instadd.get()
                ws['C'+lastx] = manuadd.get()
                ws['D'+lastx] = modeladd.get()
                ws['E'+lastx] = serialadd.get()
                ws['F'+lastx] = codenoadd.get()
                ws['G'+lastx] = intervaladd.get()
                ws['H'+lastx] = caldateadd.get()
                ws['I'+lastx] = dueadd.get()
                ws['J'+lastx] = locationadd.get()
                ws['L'+lastx] = vendoradd.get()
                ws['M'+lastx] = treatmentadd.get()
                ws['N'+lastx] = needadd.get()
                wb.save('register.xlsx')

                        
            else:
                lastx = str((ws.max_row)+1)
                ws['B'+lastx] = instadd.get()
                ws['C'+lastx] = manuadd.get()
                ws['D'+lastx] = modeladd.get()
                ws['E'+lastx] = serialadd.get()
                ws['F'+lastx] = codenoadd.get()
                ws['G'+lastx] = intervaladd.get()
                ws['H'+lastx] = caldateadd.get()
                ws['I'+lastx] = dueadd.get()
                ws['J'+lastx] = locationadd.get()
                ws['L'+lastx] = vendoradd.get()
                ws['M'+lastx] = treatmentadd.get()
                ws['N'+lastx] = needadd.get()
                wb.save('register.xlsx')

                top1.destroy()
            

        Buadd = tkinter.Button( top1, text="ADD", command=added, font=("Helevetica", 11, "bold"))
        Buadd.place(x=250, y=630)


        
        
        top1.mainloop()



bsearch=tkinter.Button(window,text="CARI",command=search,font=("Helevetica",11,"bold"))
bsearch.place(x=600, y=100, relwidth= 0.1)

asearch=tkinter.Button(window,text="ADD",command=add,font=("Helevetica",11,"bold"))

Button(window, text="KOSONGKAN", command=clear_text,
       font=("Helvetica",11,"bold")).place(x=700,y=100)


def edit():
    for i in range(2,(ws.max_row)+1):
        if((codeno.get()==ws['F'+str(i)].value)):
            Found=True
            break

        else:
            Found = False

    if(Found==True):
             
            top=Toplevel()
            top.geometry("500x600")

            A=""
            
            def nam():
                global A
                if(var1.get()==1):
                    na=codeno.get()
                    s.set(na)
                    A=na
                elif(var1.get()==0):
                    s.set("")
            
                
            def ph():
                if(var2.get()==1):
                    ph = inst.get()
                    t.set(ph)
                    B=ph
                elif(var2.get()==0):
                    t.set("")

            def ad():
                if(var3.get()==1):
                    add=model.get()
                    u.set(add)
                    C=add
                elif(var3.get()==0):
                    u.set("")

            def du():
                if(var4.get()==1):
                    add=duedate.get()
                    d.set(add)
                    D=add
                elif(var4.get()==0):
                    d.set("")
            
            def lo():
                if(var5.get() == 1):
                    add =location.get()
                    l.set(add)
                    C = add
                elif(var5.get() == 0):
                    l.set("")
            label_title1 = tkinter.Label(
                top, text="Edit Item", font=("bold", 14))
            label_title1.place(x=1, y=10)
            label_arahan2 = tkinter.Label(
                top, text="Please Fill In The Required Item to Update", font=("bold", 14))
            label_arahan2.place(x=1, y=30)
            label_codeno=tkinter.Label(top,text="New Code Number",font=("bold",14))
            label_codeno.place(x=100,y=75)

            s = StringVar()
            t = StringVar()
            u = StringVar()
            d = StringVar()
            l = StringVar()

            
            codeno1=Entry(top,textvariable=s)
            codeno1.place(x=100,y=100)

            var1 = IntVar()
            c1=Checkbutton(top, text="same as before",variable=var1,command=nam)
            c1.place(x=100,y=120)
            
            label_inst=tkinter.Label(top,text="New Instrument",font=("bold",14))
            label_inst.place(x=100,y=150)

            inst1=Entry(top,textvariable=t)
            inst1.place(x=100,y=180)

            var2 = IntVar()
            c2=Checkbutton(top, text="same as before", variable=var2,command=ph)
            c2.place(x=100,y=200)
                
            label_model=tkinter.Label(top,text="New Model Number",font=("bold",14))
            label_model.place(x=100,y=230)

            model1=Entry(top,textvariable=u)
            model1.place(x=100,y=260)

            var3 = IntVar()
            c3=Checkbutton(top, text="same as before", variable=var3,command=ad)
            c3.place(x=100,y=280)
            
            label_duedate = tkinter.Label(top, text="New Duedate", font=("bold", 14))
            label_duedate.place(x=100,y=310)

            duedate1=Entry(top,textvariable=d)
            duedate1.place(x=100,y=340)

    
            label_location = tkinter.Label(
                top, text="New Location", font=("bold", 14))
            label_location.place(x=100,y=400)

            location1 = Entry(top, textvariable=l)
            location1.place(x=100, y=430)

            var5 = IntVar()
            c5 = Checkbutton(top, text="same as before",
                             variable=var5, command=lo)
            c5.place(x=100, y=450)

            def update():
                ws['F'+str(i)] = codeno1.get()
                ws['B'+str(i)] = inst1.get()
                ws['D'+str(i)] = model1.get()
                ws['I'+str(i)] = duedate1.get()
                ws['J'+str(i)] = location1.get()
                wb.save('register.xlsx')
                
                print(A)
                
                messagebox.showinfo("Success", "Update Successfully!")
             
            
            
            Bupdate=tkinter.Button(top,text="Update",command=update,font=("Helevetica",11,"bold"))
            Bupdate.place(x=100,y=520)

            
            top.mainloop()
        
        
                
    else:
        result_add.config(text="no record found to update!")


def delete():
    for i in range(2, (ws.max_row)+1):
        if((codeno.get() == ws['F'+str(i)].value)):
            Found = True
            break

        else:
                Found = False

    if(Found == True):
        reply = messagebox.askyesno('Delete', 'Are You Sure Want To Delete This Information?')
        if reply == True :
                messagebox.showinfo('', 'Information Deleted !')
                ws.delete_rows(idx=1)
                ws.save('register.xlsx')
        else :
                        pass

    else:
        result_add.config(text="no record found to Delete!")

var1 = IntVar()
Checkbutton(window, text="no change", variable=var1)
edit=tkinter.Button(window,text="EDIT",command=edit,font=("Helevetica",11,"bold"))

dsearch=tkinter.Button(window,text="DELETE",command=delete,font=("Helevetica",11,"bold"))


asearch.place(x=280, y=180,relwidth=0.1 )
edit.place(x=380, y=180, relwidth=0.1)
dsearch.place(x=480, y=180, relwidth=0.1)


res=tkinter.Label(window,text="RESULT",font=("bold",17,"underline"))
res.place(x=400,y=230)
resbm = tkinter.Label(window, text="HASIL CARIAN", font=("bold", 17, "underline"))
resbm.place(x=365, y=270)



result_codeno=tkinter.Label(window,font=("bold",18))
result_codeno.place(x=300, y=310)

result_inst=tkinter.Label(window,font=("bold",18))
result_inst.place(x=300, y=340)

result_model=tkinter.Label(window,font=("bold",18))
result_model.place(x=300, y=370)

result_duedate = tkinter.Label(window, font=("bold", 18))
result_duedate.place(x=300, y=400)

result_location = tkinter.Label(window, font=("bold", 18))
result_location.place(x=300, y=430)

result_vendor = tkinter.Label(window, font=("bold", 18))
result_vendor.place(x=300, y=460)

result_add=tkinter.Label(window,font=("bold",18))
result_add.place(x=300, y=490)

label_codeno = tkinter.Label(window, text="Author: Sharifah Farhanim", font=1)
label_codeno.place(x=1, y=568)


window.config(menu=menubar)

window.mainloop()


